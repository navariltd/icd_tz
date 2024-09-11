# Copyright (c) 2024, Nelson Mpanju and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from openpyxl import load_workbook
from datetime import datetime

class ShipManifest(Document):
    def before_save(self):
        if self.manifest:
            file_url = self.manifest
            file_path = frappe.get_site_path('private', 'files', file_url.split('/files/')[-1])
            
            # Load the Excel file
            workbook = load_workbook(file_path, data_only=True)

            # Function to convert dates
            def convert_date(excel_date):
                if isinstance(excel_date, datetime):
                    return excel_date.strftime('%Y-%m-%d')
                if isinstance(excel_date, str):
                    try:
                        return datetime.strptime(excel_date, '%d/%m/%Y').strftime('%Y-%m-%d')
                    except ValueError:
                        return None
                return None

            # Process the VesselInformation sheet
            vessel_info_sheet = workbook['VesselInformation']
            vessel_info_row = next(vessel_info_sheet.iter_rows(min_row=2, values_only=True))
            self.mrn = vessel_info_row[0]
            self.vessel_name = vessel_info_row[1]
            self.tpa_uid = vessel_info_row[2]
            self.voyage = vessel_info_row[3]
            self.arrival_date = convert_date(vessel_info_row[4])
            self.departure_date = convert_date(vessel_info_row[5])
            self.call_sign = vessel_info_row[6]

            # Process the Containers sheet
            containers_sheet = workbook['Containers']
            for row in containers_sheet.iter_rows(min_row=2, values_only=True):
                container = self.append("containers", {})
                container.m_bl_no = row[0]
                container.type_of_container = row[1]
                container.container_no = row[2]
                container.container_size = row[3]
                container.seal_no1 = row[4]
                container.seal_no2 = row[5]
                container.seal_no3 = row[6]
                container.freight_indicator = row[7]
                container.no_of_packages = row[8]
                container.package_unit = row[9]
                container.volume = row[10]
                container.volume_unit = row[11]
                container.weight = row[12]
                container.weight_unit = row[13]
                container.plug_type_of_reefer = row[14]
                container.minimum_temperature = row[15]
                container.maximum_temperature = row[16]

            # Process the HBlContainers sheet
            hbl_containers_sheet = workbook['HBlContainers']
            for row in hbl_containers_sheet.iter_rows(min_row=2, values_only=True):
                hbicontainer = self.append("hbicontainers", {})
                hbicontainer.m_bl_no = row[0]
                hbicontainer.h_bl_no = row[1]
                hbicontainer.type_of_container = row[2]
                hbicontainer.container_no = row[3]
                hbicontainer.container_size = row[4]
                hbicontainer.seal_no1 = row[5]
                hbicontainer.seal_no2 = row[6]
                hbicontainer.seal_no3 = row[7]
                hbicontainer.freight_indicator = row[8]
                hbicontainer.no_of_packages = row[9]
                hbicontainer.package_unit = row[10]
                hbicontainer.volume = row[11]
                hbicontainer.volume_unit = row[12]
                hbicontainer.weight = row[13]
                hbicontainer.weight_unit = row[14]
                hbicontainer.plug_type_of_reefer = row[15]
                hbicontainer.minimum_temperature = row[16]
                hbicontainer.maximum_temperature = row[17]

            # Process the MasterBl sheet
            master_bl_sheet = workbook['MasterBl']
            for row in master_bl_sheet.iter_rows(min_row=2, values_only=True):
                masterbi = self.append("table_wyqi", {})
                masterbi.m_bl_no = row[0]
                masterbi.cargo_classification = row[1]
                masterbi.bl_type = row[2]
                masterbi.place_of_destination = row[3]
                masterbi.place_of_delivery = row[4]
                masterbi.oil_type = row[5]
                masterbi.port_of_loading = row[6]
                masterbi.number_of_package = row[7]
                masterbi.package_unit = row[8]  # Ensure this field is of type 'Text' in Frappe
                masterbi.gross_weight = row[9]
                masterbi.gross_weight_unit = row[10]
                masterbi.gross_volume = row[11]
                masterbi.gross_volume_unit = row[12]
                masterbi.invoice_value = row[13]
                masterbi.invoice_currency = row[14]
                masterbi.freight_charge = row[15]
                masterbi.freight_currency = row[16]
                masterbi.imdg_code = row[17]
                masterbi.packing_type = row[18]
                masterbi.shipping_agent_code = row[19]
                masterbi.shipping_agent_name = row[20]
                masterbi.forwarder_code = row[21]
                masterbi.forwarder_name = row[22]
                masterbi.forwarder_tel = row[23]
                masterbi.exporter_name = row[24]
                masterbi.exporter_tel = row[25]
                masterbi.exporter_address = row[26]
                masterbi.exporter_tin = row[27]
                masterbi.cosignee_name = row[28]
                masterbi.cosignee_tel = row[29]
                masterbi.cosignee_address = row[30]
                masterbi.cosignee_tin = row[31]
                masterbi.notify_name = row[32]
                masterbi.notify_tel = row[33]
                masterbi.notify_address = row[34]
                masterbi.notify_tin = row[35]
                masterbi.shipping_mark = row[36]
                masterbi.net_weight = row[37]
                masterbi.net_weight_unit = row[38]

            # Process the HouseBl sheet
            house_bl_sheet = workbook['HouseBl']
            for row in house_bl_sheet.iter_rows(min_row=2, values_only=True):
                housebi = self.append("housebi", {})
                housebi.m_bl_no = row[0]
                housebi.h_bl_no = row[1]
                housebi.cargo_classification = row[2]
                housebi.place_of_destination = row[3]
                housebi.net_weight = row[4]
                housebi.net_weight_unit = row[5]
                housebi.description_of_goods = row[6]
                housebi.number_of_package = row[7]
                housebi.package_unit = row[8]  # Ensure this field is of type 'Text' in Frappe
                housebi.consignee_name = row[9]
                housebi.consignee_tel = row[10]
                housebi.consignee_address = row[11]
                housebi.consignee_tin = row[12]
                housebi.notify_name = row[13]
                housebi.notify_tel = row[14]
                housebi.notify_address = row[15]
                housebi.notify_tin = row[16]
                housebi.shipping_mark = row[17]
                housebi.oil_type = row[18]
